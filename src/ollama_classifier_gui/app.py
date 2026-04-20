"""Ollama Classifier GUI - Main application.

A desktop Flet app for text classification using Ollama.
"""

import asyncio
import json
import shutil
from pathlib import Path
from typing import Any

import flet as ft
import flet_secure_storage as fss
from ollama import AsyncClient
from ollama_classifier import OllamaClassifier

from .utils import (
    check_system_dependencies,
    show_missing_dependencies,
    load_config,
    save_config,
)


class OllamaClassifierApp:
    """Main application class for Ollama Classifier GUI."""

    def __init__(self, page: ft.Page):
        self.page = page
        self.config = load_config()
        self.client: AsyncClient | None = None
        self.classifier: OllamaClassifier | None = None
        self.secure_storage = fss.SecureStorage()

        # ---- UI refs: Settings ----
        self.host_field = ft.Ref[ft.TextField]()
        self.model_dropdown = ft.Ref[ft.Dropdown]()
        self.api_key_field = ft.Ref[ft.TextField]()
        self.theme_switch = ft.Ref[ft.Switch]()
        self.test_connection_btn = ft.Ref[ft.ElevatedButton]()
        self.connection_status = ft.Ref[ft.Text]()

        # ---- UI refs: Data Input ----
        self.data_file_path_text = ft.Ref[ft.Text]()
        self.sheet_dropdown = ft.Ref[ft.Dropdown]()
        self.text_column_dropdown = ft.Ref[ft.Dropdown]()
        self.data_preview_table = ft.Ref[ft.DataTable]()

        # ---- UI refs: Schema ----
        self.manual_labels_list = ft.Ref[ft.Column]()
        self.schema_file_path_text = ft.Ref[ft.Text]()
        self.label_column_dropdown = ft.Ref[ft.Dropdown]()
        self.description_column_dropdown = ft.Ref[ft.Dropdown]()
        self.schema_preview_text = ft.Ref[ft.Text]()
        self.system_prompt_field = ft.Ref[ft.TextField]()
        self.classify_method_radio = ft.Ref[ft.RadioGroup]()
        self.save_all_probs_switch = ft.Ref[ft.Switch]()

        # ---- UI refs: Results ----
        self.results_progress = ft.Ref[ft.ProgressBar]()
        self.results_status = ft.Ref[ft.Text]()
        self.results_table = ft.Ref[ft.DataTable]()
        self.run_btn = ft.Ref[ft.ElevatedButton]()
        self.save_btn = ft.Ref[ft.ElevatedButton]()

        # ---- UI refs: Navigation ----
        self.settings_view = ft.Ref[ft.Column]()
        self.data_input_view = ft.Ref[ft.Column]()
        self.schema_view = ft.Ref[ft.Column]()
        self.results_view = ft.Ref[ft.Column]()

        # ---- App state ----
        self.data_file: str | None = None
        self.data_df: Any = None  # Polars DataFrame
        self.schema_file: str | None = None
        self.schema_df: Any = None  # Polars DataFrame for labels from file
        self.labels: dict[str, str] = {}  # label_name -> description (or empty str)
        self.results: list[dict] = []
        self._classifying = False

    # ==================================================================
    # Lifecycle
    # ==================================================================

    async def main(self):
        """Main entry point for the app."""
        missing = check_system_dependencies()
        if missing:
            await show_missing_dependencies(self.page, missing)
            await asyncio.sleep(1)
            self.page.window.destroy()
            return

        self.page.title = "Ollama Classifier GUI"
        self.page.theme_mode = self._theme_mode_from_config()
        self.page.window.width = 1200
        self.page.window.height = 800
        self.page.padding = 10

        self._build_ui()
        await self._init_ollama_client()
        await self.page.update_async()

    # ==================================================================
    # Theme helpers
    # ==================================================================

    def _theme_mode_from_config(self) -> ft.ThemeMode:
        t = self.config.get("theme", "system")
        if t == "dark":
            return ft.ThemeMode.DARK
        if t == "light":
            return ft.ThemeMode.LIGHT
        return ft.ThemeMode.SYSTEM

    # ==================================================================
    # UI construction
    # ==================================================================

    def _build_ui(self):
        self.page.add(
            ft.Row(
                controls=[
                    self._build_nav_rail(),
                    ft.VerticalDivider(width=1),
                    ft.Column(
                        controls=[
                            ft.Column(ref=self.settings_view, expand=True, visible=True, scroll=ft.ScrollMode.AUTO),
                            ft.Column(ref=self.data_input_view, expand=True, visible=False, scroll=ft.ScrollMode.AUTO),
                            ft.Column(ref=self.schema_view, expand=True, visible=False, scroll=ft.ScrollMode.AUTO),
                            ft.Column(ref=self.results_view, expand=True, visible=False, scroll=ft.ScrollMode.AUTO),
                        ],
                        expand=True,
                        scroll=ft.ScrollMode.AUTO,
                    ),
                ],
                expand=True,
            )
        )
        self._build_settings_view()
        self._build_data_input_view()
        self._build_schema_view()
        self._build_results_view()

    def _build_nav_rail(self) -> ft.NavigationRail:
        return ft.NavigationRail(
            selected_index=0,
            label_type=ft.NavigationRailLabelType.ALL,
            min_width=100,
            min_extended_width=200,
            destinations=[
                ft.NavigationRailDestination(icon=ft.Icons.SETTINGS, selected_icon=ft.Icons.SETTINGS, label="Settings"),
                ft.NavigationRailDestination(icon=ft.Icons.UPLOAD_FILE, selected_icon=ft.Icons.UPLOAD_FILE, label="Data Input"),
                ft.NavigationRailDestination(icon=ft.Icons.LABEL, selected_icon=ft.Icons.LABEL, label="Schema"),
                ft.NavigationRailDestination(icon=ft.Icons.ANALYTICS, selected_icon=ft.Icons.ANALYTICS, label="Results"),
            ],
            on_change=self._on_nav_change,
        )

    def _on_nav_change(self, e: ft.ControlEvent):
        views = [self.settings_view, self.data_input_view, self.schema_view, self.results_view]
        for i, v in enumerate(views):
            v.current.visible = (i == e.control.selected_index)
        self.page.update()

    # ---------- Settings view ----------

    def _build_settings_view(self):
        host = self.config.get("host", "http://localhost:11434")
        model = self.config.get("model", "llama3.2")
        theme = self.config.get("theme", "system")

        self.settings_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Ollama Connection", size=20, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.TextField(ref=self.host_field, label="Ollama Host URL", value=host, width=500),
                        ft.Row([
                            ft.Dropdown(ref=self.model_dropdown, label="Model", value=model, width=400,
                                        options=[ft.dropdown.Option(model)]),
                            ft.ElevatedButton("Test Connection", ref=self.test_connection_btn,
                                              icon=ft.Icons.REFRESH, on_click=self._on_test_connection),
                        ]),
                        ft.Text("", ref=self.connection_status, color=ft.Colors.GREY),
                        ft.TextField(ref=self.api_key_field, label="API Key (optional)", password=True,
                                     can_reveal_password=True, width=500,
                                     helper_text="Only needed for cloud Ollama instances"),
                    ], spacing=10),
                    padding=20,
                )
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Appearance", size=20, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.Row([
                            ft.Text("Dark Mode"),
                            ft.Switch(ref=self.theme_switch, value=(theme == "dark"),
                                      on_change=self._on_theme_change),
                        ]),
                    ], spacing=10),
                    padding=20,
                )
            ),
            ft.ElevatedButton("Save Settings", icon=ft.Icons.SAVE, on_click=self._on_save_settings),
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Desktop Shortcut", size=20, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.Text("Create a desktop shortcut to launch this app via `uvx ollama-classifier-gui`.\n"
                                "Requires `uv` to be installed on your system."),
                        ft.ElevatedButton("Create Desktop Shortcut", icon=ft.Icons.DESKTOP_WINDOWS,
                                          on_click=self._on_create_shortcut),
                    ], spacing=10),
                    padding=20,
                )
            ),
        ]

    # ---------- Data Input view ----------

    def _build_data_input_view(self):
        self.data_input_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Load Data File", size=20, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.Row([
                            ft.ElevatedButton("Select File", icon=ft.Icons.UPLOAD_FILE,
                                              on_click=self._on_select_data_file),
                            ft.Text("No file selected", ref=self.data_file_path_text, color=ft.Colors.GREY),
                        ]),
                        ft.Dropdown(ref=self.sheet_dropdown, label="Sheet (Excel only)", width=400, visible=False,
                                    on_change=self._on_sheet_change),
                        ft.Dropdown(ref=self.text_column_dropdown, label="Text Column", width=400, visible=False),
                    ], spacing=10),
                    padding=20,
                )
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Data Preview (first 5 rows)", size=18, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.DataTable(ref=self.data_preview_table, columns=[], rows=[],
                                     border=ft.border.all(1, ft.Colors.OUTLINE),
                                     horizontal_lines=ft.border.BorderSide(1, ft.Colors.OUTLINE)),
                    ], spacing=10),
                    padding=20,
                )
            ),
        ]

    # ---------- Schema view ----------

    def _build_schema_view(self):
        self.schema_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Classification Labels", size=20, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.Tabs(
                            tabs=[
                                ft.Tab(
                                    text="Manual Labels",
                                    content=ft.Column([
                                        ft.ElevatedButton("Add Label", icon=ft.Icons.ADD,
                                                          on_click=self._on_add_label),
                                        ft.Column(ref=self.manual_labels_list, spacing=5),
                                    ], spacing=10),
                                ),
                                ft.Tab(
                                    text="Load from File",
                                    content=ft.Column([
                                        ft.Row([
                                            ft.ElevatedButton("Select File", icon=ft.Icons.UPLOAD_FILE,
                                                              on_click=self._on_select_schema_file),
                                            ft.Text("No file selected", ref=self.schema_file_path_text,
                                                    color=ft.Colors.GREY),
                                        ]),
                                        ft.Dropdown(ref=self.label_column_dropdown, label="Label Column",
                                                    width=400, visible=False),
                                        ft.Dropdown(ref=self.description_column_dropdown,
                                                    label="Description Column (optional)", width=400, visible=False),
                                        ft.Text("", ref=self.schema_preview_text, selectable=True),
                                    ], spacing=10),
                                ),
                            ],
                        ),
                    ], spacing=10),
                    padding=20,
                )
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Classification Options", size=18, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.TextField(ref=self.system_prompt_field, label="Custom System Prompt (optional)",
                                     multiline=True, min_lines=3, max_lines=6,
                                     helper_text="Override the default classification prompt"),
                        ft.Text("Classification Method:"),
                        ft.RadioGroup(
                            ref=self.classify_method_radio,
                            content=ft.Column([
                                ft.Radio(value="classify", label="Classify (single call, prediction + confidence)"),
                                ft.Radio(value="score", label="Score (multi-call, all probabilities)"),
                            ]),
                            value="classify",
                            on_change=self._on_classify_method_change,
                        ),
                        ft.Row([
                            ft.Text("Save all label probabilities"),
                            ft.Switch(ref=self.save_all_probs_switch, value=False, disabled=True),
                        ]),
                    ], spacing=10),
                    padding=20,
                )
            ),
        ]

    # ---------- Results view ----------

    def _build_results_view(self):
        self.results_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column([
                        ft.Text("Classification Results", size=20, weight=ft.FontWeight.BOLD),
                        ft.Divider(),
                        ft.ProgressBar(ref=self.results_progress, visible=False, width=500),
                        ft.Text("", ref=self.results_status, color=ft.Colors.GREY),
                        ft.DataTable(ref=self.results_table, columns=[], rows=[],
                                     border=ft.border.all(1, ft.Colors.OUTLINE),
                                     horizontal_lines=ft.border.BorderSide(1, ft.Colors.OUTLINE)),
                    ], spacing=10),
                    padding=20,
                )
            ),
            ft.Row([
                ft.ElevatedButton("Run Classification", ref=self.run_btn, icon=ft.Icons.PLAY_ARROW,
                                  on_click=self._on_run_classification),
                ft.ElevatedButton("Save Results", ref=self.save_btn, icon=ft.Icons.SAVE,
                                  on_click=self._on_save_results, disabled=True),
            ]),
        ]

    # ==================================================================
    # Ollama client
    # ==================================================================

    async def _init_ollama_client(self):
        host = self.config.get("host", "http://localhost:11434")
        api_key = await self.secure_storage.get("api_key")
        headers: dict[str, str] = {}
        if api_key:
            headers["Authorization"] = f"Bearer {api_key}"
        self.client = AsyncClient(host=host, headers=headers)
        model = self.config.get("model", "llama3.2")
        self.classifier = OllamaClassifier(self.client, model=model)

    # ==================================================================
    # Settings handlers
    # ==================================================================

    async def _on_test_connection(self, e: ft.ControlEvent):
        self.test_connection_btn.current.disabled = True
        self.connection_status.current.value = "Testing connection..."
        self.connection_status.current.color = ft.Colors.GREY
        await self.page.update_async()
        try:
            host = self.host_field.current.value
            api_key = self.api_key_field.current.value
            headers: dict[str, str] = {}
            if api_key:
                headers["Authorization"] = f"Bearer {api_key}"
            tmp_client = AsyncClient(host=host, headers=headers)
            models_resp = await tmp_client.list()
            model_names = sorted(m["name"] for m in models_resp.get("models", []))
            self.model_dropdown.current.options = [ft.dropdown.Option(n) for n in model_names]
            if model_names:
                self.model_dropdown.current.value = model_names[0]
            self.connection_status.current.value = f"✓ Connected — {len(model_names)} model(s) found."
            self.connection_status.current.color = ft.Colors.GREEN
        except Exception as ex:
            self.connection_status.current.value = f"✗ Error: {ex}"
            self.connection_status.current.color = ft.Colors.RED
        finally:
            self.test_connection_btn.current.disabled = False
            await self.page.update_async()

    async def _on_theme_change(self, e: ft.ControlEvent):
        is_dark = e.control.value
        self.config["theme"] = "dark" if is_dark else "light"
        self.page.theme_mode = ft.ThemeMode.DARK if is_dark else ft.ThemeMode.LIGHT
        save_config(self.config)
        await self.page.update_async()

    async def _on_save_settings(self, e: ft.ControlEvent):
        self.config["host"] = self.host_field.current.value
        self.config["model"] = self.model_dropdown.current.value
        save_config(self.config)
        api_key = self.api_key_field.current.value
        if api_key:
            await self.secure_storage.set("api_key", api_key)
        else:
            try:
                await self.secure_storage.remove("api_key")
            except Exception:
                pass
        await self._init_ollama_client()
        await self._show_dialog("Settings Saved", "Your settings have been saved successfully.")

    async def _on_create_shortcut(self, e: ft.ControlEvent):
        if shutil.which("uv") is None:
            await self._show_dialog(
                "uv Not Found",
                "The `uv` package manager is required to create a shortcut.\n\n"
                "Install it from https://docs.astral.sh/uv/getting-started/installation/",
            )
            return
        try:
            from pyshortcuts import make_shortcut
            make_shortcut(
                script="uvx",
                arguments="ollama-classifier-gui",
                name="Ollama Classifier GUI",
                terminal=False,
                desktop=True,
                startmenu=True,
            )
            await self._show_dialog("Shortcut Created",
                                    "Desktop and Start Menu shortcuts have been created.")
        except Exception as ex:
            await self._show_dialog("Error Creating Shortcut", str(ex))

    # ==================================================================
    # Data Input handlers
    # ==================================================================

    async def _on_select_data_file(self, e: ft.ControlEvent):
        import polars as pl

        result = await self._pick_file(
            allowed_extensions=["csv", "xlsx", "xls", "json"],
            dialog_title="Select data file",
        )
        if not result:
            return
        file_path = result
        self.data_file = file_path
        self.data_file_path_text.current.value = file_path
        self.sheet_dropdown.current.visible = False
        self.text_column_dropdown.current.visible = False

        try:
            if file_path.endswith((".xlsx", ".xls")):
                import openpyxl
                wb = openpyxl.load_workbook(file_path, read_only=True)
                sheet_names = wb.sheetnames
                wb.close()
                if not sheet_names:
                    raise ValueError("Excel file has no sheets.")
                self.sheet_dropdown.current.options = [ft.dropdown.Option(n) for n in sheet_names]
                self.sheet_dropdown.current.value = sheet_names[0]
                self.sheet_dropdown.current.visible = True
                self.data_df = pl.read_excel(file_path, engine="openpyxl", sheet_name=sheet_names[0])
            elif file_path.endswith(".csv"):
                self.data_df = pl.read_csv(file_path)
            elif file_path.endswith(".json"):
                try:
                    self.data_df = pl.read_json(file_path)
                except Exception:
                    # Fallback: try reading raw JSON and normalising
                    self.data_df = self._try_flatten_json(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_path}")

            if self.data_df is not None and len(self.data_df) > 0:
                cols = self.data_df.columns
                self.text_column_dropdown.current.options = [ft.dropdown.Option(c) for c in cols]
                self.text_column_dropdown.current.value = cols[0]
                self.text_column_dropdown.current.visible = True
                self._refresh_data_preview()
            else:
                raise ValueError("The file contains no data rows.")

        except Exception as ex:
            await self._show_dialog("Error Loading File", str(ex))

        await self.page.update_async()

    async def _on_sheet_change(self, e: ft.ControlEvent):
        import polars as pl

        sheet_name = self.sheet_dropdown.current.value
        if not sheet_name or not self.data_file:
            return
        try:
            self.data_df = pl.read_excel(self.data_file, engine="openpyxl", sheet_name=sheet_name)
            cols = self.data_df.columns
            self.text_column_dropdown.current.options = [ft.dropdown.Option(c) for c in cols]
            self.text_column_dropdown.current.value = cols[0]
            self._refresh_data_preview()
        except Exception as ex:
            await self._show_dialog("Error Loading Sheet", str(ex))
        await self.page.update_async()

    @staticmethod
    def _try_flatten_json(file_path: str):
        """Attempt to flatten a JSON file into a Polars DataFrame."""
        import polars as pl

        with open(file_path, "r", encoding="utf-8") as f:
            raw = json.load(f)

        # If it's a list of dicts
        if isinstance(raw, list):
            # Try to flatten each record
            flat_records = []
            for rec in raw:
                if isinstance(rec, dict):
                    flat_records.append(pl.DataFrame(rec).unnest(pl.all()))
                else:
                    flat_records.append({"value": rec})
            return pl.concat(flat_records, how="diagonal") if flat_records else pl.DataFrame()

        # If it's a single dict
        if isinstance(raw, dict):
            # Check if any value is a nested structure
            has_nested = any(isinstance(v, (list, dict)) for v in raw.values())
            if has_nested:
                raise ValueError(
                    "JSON contains nested structures that cannot be flattened to a simple table. "
                    "Please provide a flat JSON (list of objects with primitive values)."
                )
            return pl.DataFrame(raw)

        raise ValueError("Unsupported JSON structure. Expected a list of objects or a flat object.")

    def _refresh_data_preview(self):
        if self.data_df is None:
            return
        preview = self.data_df.head(5)
        self.data_preview_table.current.columns = [ft.DataColumn(ft.Text(c)) for c in preview.columns]
        self.data_preview_table.current.rows = []
        for row in preview.iter_rows(named=True):
            self.data_preview_table.current.rows.append(
                ft.DataRow(cells=[ft.DataCell(ft.Text(str(row[c]))) for c in preview.columns])
            )

    # ==================================================================
    # Schema handlers
    # ==================================================================

    async def _on_add_label(self, e: ft.ControlEvent):
        row = ft.Row([
            ft.TextField(label="Label", width=200, expand=True),
            ft.TextField(label="Description (optional)", width=300, expand=True),
            ft.IconButton(icon=ft.Icons.DELETE,
                          on_click=lambda ev, r=row: self._on_remove_label(r)),
        ], spacing=10)
        self.manual_labels_list.current.controls.append(row)
        await self.page.update_async()

    async def _on_remove_label(self, row: ft.Row):
        self.manual_labels_list.current.controls.remove(row)
        await self.page.update_async()

    async def _on_select_schema_file(self, e: ft.ControlEvent):
        import polars as pl

        result = await self._pick_file(
            allowed_extensions=["csv", "xlsx", "xls", "json"],
            dialog_title="Select schema / labels file",
        )
        if not result:
            return
        file_path = result
        self.schema_file = file_path
        self.schema_file_path_text.current.value = file_path

        try:
            if file_path.endswith((".xlsx", ".xls")):
                self.schema_df = pl.read_excel(file_path, engine="openpyxl")
            elif file_path.endswith(".csv"):
                self.schema_df = pl.read_csv(file_path)
            elif file_path.endswith(".json"):
                try:
                    self.schema_df = pl.read_json(file_path)
                except Exception:
                    self.schema_df = self._try_flatten_json(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_path}")

            cols = self.schema_df.columns
            self.label_column_dropdown.current.options = [ft.dropdown.Option(c) for c in cols]
            self.description_column_dropdown.current.options = [ft.dropdown.Option(c) for c in cols]
            self.label_column_dropdown.current.value = cols[0]
            self.label_column_dropdown.current.visible = True
            self.description_column_dropdown.current.visible = True
            self._refresh_schema_preview()

        except Exception as ex:
            await self._show_dialog("Error Loading Schema File", str(ex))
        await self.page.update_async()

    def _refresh_schema_preview(self):
        if self.schema_df is None:
            return
        label_col = self.label_column_dropdown.current.value
        desc_col = self.description_column_dropdown.current.value
        preview_rows = self.schema_df.head(10)
        lines = []
        for row in preview_rows.iter_rows(named=True):
            label = str(row.get(label_col, ""))
            desc = str(row.get(desc_col, "")) if desc_col and desc_col != label_col else ""
            lines.append(f"• {label}" + (f" — {desc}" if desc else ""))
        self.schema_preview_text.current.value = "\n".join(lines) if lines else "No labels found."

    def _collect_labels(self) -> dict[str, str]:
        """Collect labels from either manual input or schema file."""
        import polars as pl

        active_tab = self.schema_view.current.controls[0].content.controls[2].selected_index  # type: ignore[union-attr]

        if active_tab == 0:
            # Manual labels
            labels: dict[str, str] = {}
            for row_ctrl in self.manual_labels_list.current.controls:
                fields = [c for c in row_ctrl.controls if isinstance(c, ft.TextField)]
                if len(fields) >= 1 and fields[0].value and fields[0].value.strip():
                    label_name = fields[0].value.strip()
                    desc = fields[1].value.strip() if len(fields) >= 2 and fields[1].value else ""
                    labels[label_name] = desc
            return labels
        else:
            # From file
            if self.schema_df is None:
                return {}
            label_col = self.label_column_dropdown.current.value
            desc_col = self.description_column_dropdown.current.value
            labels = {}
            for row in self.schema_df.iter_rows(named=True):
                name = str(row.get(label_col, "")).strip()
                if not name:
                    continue
                desc = ""
                if desc_col and desc_col != label_col:
                    desc = str(row.get(desc_col, "")).strip()
                labels[name] = desc
            return labels

    # ==================================================================
    # Classification
    # ==================================================================

    async def _on_classify_method_change(self, e: ft.ControlEvent):
        method = e.control.value
        self.save_all_probs_switch.current.disabled = (method != "score")
        if method != "score":
            self.save_all_probs_switch.current.value = False
        await self.page.update_async()

    async def _on_run_classification(self, e: ft.ControlEvent):
        if self._classifying:
            return

        # --- Validate inputs ---
        if self.data_df is None:
            await self._show_dialog("No Data", "Please load a data file first (Data Input tab).")
            return

        text_col = self.text_column_dropdown.current.value
        if not text_col:
            await self._show_dialog("No Column", "Please select a text column.")
            return

        labels = self._collect_labels()
        if not labels:
            await self._show_dialog("No Labels", "Please define at least one classification label (Schema tab).")
            return

        # Build choices: if all descriptions are empty, use list form
        all_empty_desc = all(d == "" for d in labels.values())
        choices = list(labels.keys()) if all_empty_desc else labels

        method = self.classify_method_radio.current.value
        save_all = self.save_all_probs_switch.current.value and method == "score"
        system_prompt = self.system_prompt_field.current.value or None

        # Reinitialise classifier with current model
        if self.classifier is None:
            await self._show_dialog("Not Connected", "Please configure and test the Ollama connection in Settings.")
            return

        # Extract texts
        texts = self.data_df[text_col].to_list()

        # --- Run ---
        self._classifying = True
        self.run_btn.current.disabled = True
        self.save_btn.current.disabled = True
        self.results_progress.current.visible = True
        self.results_progress.current.value = 0
        self.results_status.current.value = f"Starting classification of {len(texts)} item(s)…"
        self.results = []
        self.results_table.current.rows = []
        await self.page.update_async()

        classify_fn = self.classifier.ascore if method == "score" else self.classifier.aclassify

        try:
            for i, text in enumerate(texts):
                text_str = str(text)
                self.results_status.current.value = f"Classifying {i + 1}/{len(texts)}…"
                self.results_progress.current.value = (i + 1) / len(texts)
                await self.page.update_async()

                try:
                    result = await classify_fn(
                        text=text_str,
                        choices=choices,
                        system_prompt=system_prompt,
                    )
                    row_dict: dict[str, Any] = {
                        "text": text_str,
                        "prediction": result.prediction,
                        "confidence": result.confidence,
                    }
                    if save_all:
                        row_dict["probabilities"] = result.probabilities
                    self.results.append(row_dict)

                    # Add row to table
                    cells = [
                        ft.DataCell(ft.Text(text_str[:80] + ("…" if len(text_str) > 80 else ""))),
                        ft.DataCell(ft.Text(result.prediction)),
                        ft.DataCell(ft.Text(f"{result.confidence:.2%}")),
                    ]
                    self.results_table.current.rows.append(ft.DataRow(cells=cells))

                except Exception as item_ex:
                    self.results.append({
                        "text": text_str,
                        "prediction": f"ERROR: {item_ex}",
                        "confidence": 0.0,
                    })
                    self.results_table.current.rows.append(
                        ft.DataRow(cells=[
                            ft.DataCell(ft.Text(text_str[:80])),
                            ft.DataCell(ft.Text(f"ERROR", color=ft.Colors.RED)),
                            ft.DataCell(ft.Text("0.00%")),
                        ])
                    )

                # Update table header columns on first successful row
                if i == 0:
                    self.results_table.current.columns = [
                        ft.DataColumn(ft.Text("Text")),
                        ft.DataColumn(ft.Text("Prediction")),
                        ft.DataColumn(ft.Text("Confidence")),
                    ]

                await self.page.update_async()

            self.results_status.current.value = f"✓ Done — classified {len(texts)} item(s)."
            self.results_status.current.color = ft.Colors.GREEN
            self.save_btn.current.disabled = False

        except Exception as ex:
            self.results_status.current.value = f"✗ Error: {ex}"
            self.results_status.current.color = ft.Colors.RED
        finally:
            self._classifying = False
            self.run_btn.current.disabled = False
            self.results_progress.current.visible = False
            await self.page.update_async()

    # ==================================================================
    # Save results
    # ==================================================================

    async def _on_save_results(self, e: ft.ControlEvent):
        if not self.results:
            await self._show_dialog("No Results", "Nothing to save — run classification first.")
            return

        save_dir = await self._pick_save_dir(dialog_title="Select save location")
        if not save_dir:
            return

        base_path = Path(save_dir) / "classification_results"

        try:
            import polars as pl

            rows = []
            for r in self.results:
                row: dict[str, Any] = {
                    "text": r["text"],
                    "prediction": r["prediction"],
                    "confidence": r["confidence"],
                }
                if "probabilities" in r and isinstance(r["probabilities"], dict):
                    for k, v in r["probabilities"].items():
                        row[f"prob_{k}"] = v
                rows.append(row)

            df = pl.DataFrame(rows)
            csv_path = str(base_path) + ".csv"
            json_path = str(base_path) + ".json"

            df.write_csv(csv_path)
            df.write_json(json_path)

            await self._show_dialog(
                "Results Saved",
                f"CSV: {csv_path}\nJSON: {json_path}",
            )
        except Exception as ex:
            await self._show_dialog("Error Saving Results", str(ex))

    # ==================================================================
    # Helpers
    # ==================================================================

    async def _pick_file(self, allowed_extensions: list[str] | None = None,
                         dialog_title: str = "Select file") -> str | None:
        fp = ft.FilePicker()
        self.page.overlay.append(fp)
        await self.page.update_async()
        result = await fp.pick_file(
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=allowed_extensions or [],
            dialog_title=dialog_title,
        )
        # Clean up
        self.page.overlay.remove(fp)
        await self.page.update_async()
        if result and result.files:
            return result.files[0].path
        return None

    async def _pick_save_dir(self, dialog_title: str = "Select folder") -> str | None:
        fp = ft.FilePicker()
        self.page.overlay.append(fp)
        await self.page.update_async()
        result = await fp.get_directory_path(dialog_title=dialog_title)
        self.page.overlay.remove(fp)
        await self.page.update_async()
        return result

    async def _show_dialog(self, title: str, message: str):
        dlg = ft.AlertDialog(
            title=ft.Text(title),
            content=ft.Text(message, selectable=True),
            actions=[ft.TextButton("OK", on_click=lambda e: self.page.close(dlg))],
        )
        self.page.overlay.append(dlg)
        await self.page.update_async()
        await dlg.wait_dismiss()


# ======================================================================
# Entry point
# ======================================================================

def main():
    """Entry point for the Flet app (desktop mode)."""
    ft.app(target=OllamaClassifierApp)


if __name__ == "__main__":
    main()
