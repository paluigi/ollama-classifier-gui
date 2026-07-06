"""Ollama Classifier GUI - Main application.

A desktop Flet app for text classification using multiple LLM backends
(Ollama, vLLM, SGLang, llama.cpp).
"""

import asyncio
import shutil
from typing import Any

import flet as ft
import flet_secure_storage as fss

from .utils import (
    BACKEND_TYPES,
    DEFAULT_ENDPOINTS,
    app_repository_url,
    app_version,
    check_system_dependencies,
    load_config,
    save_config,
    show_missing_dependencies,
)


class OllamaClassifierApp:
    """Main application class for Ollama Classifier GUI."""

    def __init__(self, page: ft.Page):
        self.page = page
        self.config = load_config()
        self.secure_storage = fss.SecureStorage()

        # Backend state
        self._classifier: Any | None = None

        # ---- UI refs: Settings ----
        self.backend_dropdown = ft.Ref[ft.Dropdown]()
        self.endpoint_field = ft.Ref[ft.TextField]()
        self.model_field = ft.Ref[ft.TextField]()
        self.api_key_field = ft.Ref[ft.TextField]()
        self.theme_switch = ft.Ref[ft.Switch]()
        self.test_connection_btn = ft.Ref[ft.Button]()
        self.connection_status = ft.Ref[ft.Text]()
        self.batch_size_field = ft.Ref[ft.TextField]()

        # ---- UI refs: Data Input ----
        self.data_file_path_text = ft.Ref[ft.Text]()
        self.data_sheet_dropdown = ft.Ref[ft.Dropdown]()
        self.text_column_dropdown = ft.Ref[ft.Dropdown]()
        self.data_preview_table = ft.Ref[ft.DataTable]()

        # ---- UI refs: Schema ----
        self.manual_labels_list = ft.Ref[ft.Column]()
        self.schema_file_path_text = ft.Ref[ft.Text]()
        self.schema_sheet_dropdown = ft.Ref[ft.Dropdown]()
        self.label_column_dropdown = ft.Ref[ft.Dropdown]()
        self.description_column_dropdown = ft.Ref[ft.Dropdown]()
        self.schema_preview_text = ft.Ref[ft.Text]()
        self.system_prompt_field = ft.Ref[ft.TextField]()
        self.classify_method_radio = ft.Ref[ft.RadioGroup]()
        self.output_format_radio = ft.Ref[ft.RadioGroup]()
        self.schema_tabs = ft.Ref[ft.Tabs]()

        # ---- UI refs: Results ----
        self.results_progress = ft.Ref[ft.ProgressBar]()
        self.results_status = ft.Ref[ft.Text]()
        self.results_table = ft.Ref[ft.DataTable]()
        self.run_btn = ft.Ref[ft.Button]()
        self.save_btn = ft.Ref[ft.Button]()

        # ---- UI refs: Navigation ----
        self.settings_view = ft.Ref[ft.Column]()
        self.data_input_view = ft.Ref[ft.Column]()
        self.schema_view = ft.Ref[ft.Column]()
        self.results_view = ft.Ref[ft.Column]()
        self.info_view = ft.Ref[ft.Column]()

        # ---- App state ----
        self.data_file: str | None = None
        self.data_df: Any = None  # Polars DataFrame
        self.schema_file: str | None = None
        self.schema_df: Any = None
        self.labels: dict[str, str] = {}
        self.results: list[dict] = []
        self._classifying = False

    # ==================================================================
    # Lifecycle
    # ==================================================================

    async def main(self):
        """Main entry point for the app."""
        missing = check_system_dependencies()
        if missing:
            await show_missing_dependencies(self.page, missing)
            await asyncio.sleep(1)
            await self.page.window.destroy()
            return

        self.page.title = "LLM Classifier GUI"
        self.page.theme_mode = self._theme_mode_from_config()
        self.page.window.width = 1200
        self.page.window.height = 800
        self.page.padding = 10

        self._build_ui()
        self.page.update()

    # ==================================================================
    # Theme helpers
    # ==================================================================

    def _theme_mode_from_config(self) -> ft.ThemeMode:
        t = self.config.get("theme", "system")
        if t == "dark":
            return ft.ThemeMode.DARK
        if t == "light":
            return ft.ThemeMode.LIGHT
        return ft.ThemeMode.SYSTEM

    # ==================================================================
    # UI construction
    # ==================================================================

    def _build_ui(self):
        self.page.add(
            ft.Row(
                controls=[
                    self._build_nav_rail(),
                    ft.VerticalDivider(width=1),
                    ft.Column(
                        controls=[
                            ft.Column(
                                ref=self.settings_view,
                                expand=True,
                                visible=True,
                                scroll=ft.ScrollMode.AUTO,
                            ),
                            ft.Column(
                                ref=self.data_input_view,
                                expand=True,
                                visible=False,
                                scroll=ft.ScrollMode.AUTO,
                            ),
                            ft.Column(
                                ref=self.schema_view,
                                expand=True,
                                visible=False,
                                scroll=ft.ScrollMode.AUTO,
                            ),
                            ft.Column(
                                ref=self.results_view,
                                expand=True,
                                visible=False,
                                scroll=ft.ScrollMode.AUTO,
                            ),
                            ft.Column(
                                ref=self.info_view,
                                expand=True,
                                visible=False,
                                scroll=ft.ScrollMode.AUTO,
                            ),
                        ],
                        expand=True,
                        scroll=ft.ScrollMode.AUTO,
                    ),
                ],
                expand=True,
            )
        )
        self._build_settings_view()
        self._build_data_input_view()
        self._build_schema_view()
        self._build_results_view()
        self._build_info_view()

    def _build_nav_rail(self) -> ft.NavigationRail:
        return ft.NavigationRail(
            selected_index=0,
            label_type=ft.NavigationRailLabelType.ALL,
            min_width=100,
            min_extended_width=200,
            destinations=[
                ft.NavigationRailDestination(
                    icon=ft.Icons.SETTINGS,
                    selected_icon=ft.Icons.SETTINGS,
                    label="Settings",
                ),
                ft.NavigationRailDestination(
                    icon=ft.Icons.UPLOAD_FILE,
                    selected_icon=ft.Icons.UPLOAD_FILE,
                    label="Data Input",
                ),
                ft.NavigationRailDestination(
                    icon=ft.Icons.LABEL, selected_icon=ft.Icons.LABEL, label="Schema"
                ),
                ft.NavigationRailDestination(
                    icon=ft.Icons.ANALYTICS,
                    selected_icon=ft.Icons.ANALYTICS,
                    label="Results",
                ),
                ft.NavigationRailDestination(
                    icon=ft.Icons.INFO_OUTLINE,
                    selected_icon=ft.Icons.INFO,
                    label="Info",
                ),
            ],
            on_change=self._on_nav_change,
        )

    def _on_nav_change(self, e: ft.ControlEvent):
        views = [
            self.settings_view,
            self.data_input_view,
            self.schema_view,
            self.results_view,
            self.info_view,
        ]
        for i, v in enumerate(views):
            v.current.visible = i == e.control.selected_index
        self.page.update()

    # ---------- Settings view ----------

    def _build_settings_view(self):
        backend = self.config.get("backend_type", "ollama")
        endpoint = self.config.get("endpoint", DEFAULT_ENDPOINTS[backend])
        model = self.config.get("model", "llama3.2")
        theme = self.config.get("theme", "system")
        batch_size = self.config.get("batch_size", "1")

        self.settings_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Inference Backend", size=20, weight=ft.FontWeight.BOLD
                            ),
                            ft.Divider(),
                            ft.Dropdown(
                                ref=self.backend_dropdown,
                                label="Backend Type",
                                value=backend,
                                width=400,
                                options=[ft.dropdown.Option(b) for b in BACKEND_TYPES],
                                on_select=self._on_backend_type_change,
                            ),
                            ft.TextField(
                                ref=self.endpoint_field,
                                label="Endpoint URL",
                                value=endpoint,
                                width=500,
                                helper="Base URL of the inference server",
                            ),
                            ft.TextField(
                                ref=self.model_field,
                                label="Model",
                                value=model,
                                width=400,
                                helper="Model identifier (e.g. llama3.2, meta-llama/Llama-3.2-3B-Instruct)",
                            ),
                            ft.Row(
                                [
                                    ft.Button(
                                        "Test Connection",
                                        ref=self.test_connection_btn,
                                        icon=ft.Icons.REFRESH,
                                        on_click=self._on_test_connection,
                                    ),
                                    ft.Text(
                                        "",
                                        ref=self.connection_status,
                                        color=ft.Colors.GREY,
                                    ),
                                ]
                            ),
                            ft.TextField(
                                ref=self.api_key_field,
                                label="API Key (optional)",
                                password=True,
                                can_reveal_password=True,
                                width=500,
                                helper="Only needed for authenticated remote servers",
                            ),
                            ft.TextField(
                                ref=self.batch_size_field,
                                label="Batch Size",
                                value=batch_size,
                                width=200,
                                helper="Number of items per batch (default: 1)",
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text("Appearance", size=20, weight=ft.FontWeight.BOLD),
                            ft.Divider(),
                            ft.Row(
                                [
                                    ft.Text("Dark Mode"),
                                    ft.Switch(
                                        ref=self.theme_switch,
                                        value=(theme == "dark"),
                                        on_change=self._on_theme_change,
                                    ),
                                ]
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
            ft.Button(
                "Save Settings", icon=ft.Icons.SAVE, on_click=self._on_save_settings
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Desktop Shortcut", size=20, weight=ft.FontWeight.BOLD
                            ),
                            ft.Divider(),
                            ft.Text(
                                "Create a desktop shortcut to launch this app via `uvx ollama-classifier-gui`.\n"
                                "Requires `uv` to be installed on your system."
                            ),
                            ft.Button(
                                "Create Desktop Shortcut",
                                icon=ft.Icons.DESKTOP_WINDOWS,
                                on_click=self._on_create_shortcut,
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
        ]

    # ---------- Data Input view ----------

    def _build_data_input_view(self):
        self.data_input_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Load Data File", size=20, weight=ft.FontWeight.BOLD
                            ),
                            ft.Divider(),
                            ft.Row(
                                [
                                    ft.Button(
                                        "Select File",
                                        icon=ft.Icons.UPLOAD_FILE,
                                        on_click=self._on_select_data_file,
                                    ),
                                    ft.Text(
                                        "No file selected",
                                        ref=self.data_file_path_text,
                                        color=ft.Colors.GREY,
                                    ),
                                ]
                            ),
                            ft.Dropdown(
                                ref=self.data_sheet_dropdown,
                                label="Sheet (Excel only)",
                                width=400,
                                visible=False,
                                on_select=self._on_data_sheet_change,
                            ),
                            ft.Dropdown(
                                ref=self.text_column_dropdown,
                                label="Text Column",
                                width=400,
                                visible=False,
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Data Preview (first 5 rows)",
                                size=18,
                                weight=ft.FontWeight.BOLD,
                            ),
                            ft.Divider(),
                            ft.DataTable(
                                ref=self.data_preview_table,
                                columns=[ft.DataColumn(ft.Text(""))],
                                rows=[],
                                border=ft.Border.all(1, ft.Colors.OUTLINE),
                                horizontal_lines=ft.BorderSide(1, ft.Colors.OUTLINE),
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
        ]

    # ---------- Schema view ----------

    def _build_schema_view(self):
        self.schema_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Classification Labels",
                                size=20,
                                weight=ft.FontWeight.BOLD,
                            ),
                            ft.Divider(),
                            ft.Container(
                                content=ft.Tabs(
                                    ref=self.schema_tabs,
                                    selected_index=0,
                                    length=2,
                                    on_change=self._on_schema_tab_change,
                                    content=ft.Column(
                                        [
                                            ft.TabBar(
                                                tabs=[
                                                    ft.Tab(label="Manual Labels"),
                                                    ft.Tab(label="Load from File"),
                                                ]
                                            ),
                                            ft.TabBarView(
                                                controls=[
                                                    ft.Column(
                                                        [
                                                            ft.Button(
                                                                "Add Label",
                                                                icon=ft.Icons.ADD,
                                                                on_click=self._on_add_label,
                                                            ),
                                                            ft.Column(
                                                                ref=self.manual_labels_list,
                                                                spacing=5,
                                                            ),
                                                        ],
                                                        spacing=10,
                                                    ),
                                                    ft.Column(
                                                        [
                                                            ft.Row(
                                                                [
                                                                    ft.Button(
                                                                        "Select File",
                                                                        icon=ft.Icons.UPLOAD_FILE,
                                                                        on_click=self._on_select_schema_file,
                                                                    ),
                                                                    ft.Text(
                                                                        "No file selected",
                                                                        ref=self.schema_file_path_text,
                                                                        color=ft.Colors.GREY,
                                                                    ),
                                                                ]
                                                            ),
                                                            ft.Dropdown(
                                                                ref=self.schema_sheet_dropdown,
                                                                label="Sheet (Excel only)",
                                                                width=400,
                                                                visible=False,
                                                                on_select=self._on_schema_sheet_change,
                                                            ),
                                                            ft.Dropdown(
                                                                ref=self.label_column_dropdown,
                                                                label="Label Column",
                                                                width=400,
                                                                visible=False,
                                                            ),
                                                            ft.Dropdown(
                                                                ref=self.description_column_dropdown,
                                                                label="Description Column (optional)",
                                                                width=400,
                                                                visible=False,
                                                            ),
                                                            ft.Text(
                                                                "",
                                                                ref=self.schema_preview_text,
                                                                selectable=True,
                                                            ),
                                                        ],
                                                        spacing=10,
                                                    ),
                                                ],
                                                expand=True,
                                            ),
                                        ],
                                        expand=True,
                                    ),
                                ),
                                height=440,
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Classification Options",
                                size=18,
                                weight=ft.FontWeight.BOLD,
                            ),
                            ft.Divider(),
                            ft.TextField(
                                ref=self.system_prompt_field,
                                label="Custom System Prompt (optional)",
                                multiline=True,
                                min_lines=3,
                                max_lines=6,
                                helper="Override the default classification prompt",
                            ),
                            ft.Text(
                                "Classification Method:", weight=ft.FontWeight.W_500
                            ),
                            ft.RadioGroup(
                                ref=self.classify_method_radio,
                                content=ft.Column(
                                    [
                                        ft.Radio(
                                            value="classify",
                                            label="Classify (multi-call, exact confidence — slower)",
                                        ),
                                        ft.Radio(
                                            value="generate",
                                            label="Generate (adaptive, approximate confidence — faster)",
                                        ),
                                    ]
                                ),
                                value="classify",
                            ),
                            ft.Text(
                                "Output Format:",
                                weight=ft.FontWeight.W_500,
                                margin=ft.Margin.only(top=10),
                            ),
                            ft.RadioGroup(
                                ref=self.output_format_radio,
                                content=ft.Column(
                                    [
                                        ft.Radio(
                                            value="top_label",
                                            label="Top label only (prediction + confidence)",
                                        ),
                                        ft.Radio(
                                            value="all_labels",
                                            label="All labels (each label becomes a column with its probability)",
                                        ),
                                    ]
                                ),
                                value="top_label",
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
        ]

    # ---------- Results view ----------

    def _build_results_view(self):
        self.results_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column(
                        [
                            ft.Text(
                                "Classification Results",
                                size=20,
                                weight=ft.FontWeight.BOLD,
                            ),
                            ft.Divider(),
                            ft.ProgressBar(
                                ref=self.results_progress, visible=False, width=500
                            ),
                            ft.Text("", ref=self.results_status, color=ft.Colors.GREY),
                            ft.Text("Preview (first 20 rows):", size=14, italic=True),
                            ft.Container(
                                content=ft.Column(
                                    controls=[
                                        ft.DataTable(
                                            ref=self.results_table,
                                            columns=[ft.DataColumn(ft.Text(""))],
                                            rows=[],
                                            border=ft.Border.all(1, ft.Colors.OUTLINE),
                                            horizontal_lines=ft.BorderSide(
                                                1, ft.Colors.OUTLINE
                                            ),
                                        )
                                    ],
                                    scroll=ft.ScrollMode.AUTO,
                                ),
                                height=350,
                            ),
                        ],
                        spacing=10,
                    ),
                    padding=20,
                )
            ),
            ft.Row(
                [
                    ft.Button(
                        "Run Classification",
                        ref=self.run_btn,
                        icon=ft.Icons.PLAY_ARROW,
                        on_click=self._on_run_classification,
                    ),
                    ft.Button(
                        "Save Results",
                        ref=self.save_btn,
                        icon=ft.Icons.SAVE,
                        on_click=self._on_save_results,
                        disabled=True,
                    ),
                ]
            ),
        ]

    # ---------- Info view ----------

    def _build_info_view(self):
        version = app_version()
        repo_url = app_repository_url()
        controls = [
            ft.Text("LLM Classifier GUI", size=24, weight=ft.FontWeight.BOLD),
            ft.Text(
                f"Version {version}",
                size=14,
                color=ft.Colors.GREY,
            ),
        ]
        if repo_url:
            controls.append(
                ft.Column(
                    [
                        ft.Markdown(
                            f"**GitHub Repository:** [{repo_url}]({repo_url})",
                            on_tap_link=lambda e: self.page.launch_url(repo_url),
                        ),
                        ft.Button(
                            "Open in Browser",
                            icon=ft.Icons.OPEN_IN_NEW,
                            on_click=lambda e: self.page.launch_url(repo_url),
                        )
                    ],
                    spacing=5,
                )
            )

        self.info_view.current.controls = [
            ft.Card(
                content=ft.Container(
                    content=ft.Column(controls, spacing=10),
                    padding=20,
                )
            )
        ]

    # ==================================================================
    # Backend / classifier initialisation
    # ==================================================================

    async def _get_classifier(self):
        """Return the appropriate classifier based on current settings."""
        backend_type = self.config.get("backend_type", "ollama")
        endpoint = self.config.get("endpoint", DEFAULT_ENDPOINTS[backend_type])
        model = self.config.get("model", "llama3.2")
        api_key = await self.secure_storage.get("api_key") or ""

        backend = self._create_backend(backend_type, endpoint, model, api_key)
        from ollama_classifier import LLMClassifier

        self._classifier = LLMClassifier(backend)
        return self._classifier

    def _create_backend(
        self, backend_type: str, endpoint: str, model: str, api_key: str
    ):
        """Create the appropriate backend based on backend_type."""
        if backend_type == "ollama":
            from ollama_classifier.backends import OllamaBackend

            return OllamaBackend(model=model, host=endpoint)
        elif backend_type == "vllm":
            from ollama_classifier.backends import VLLMBackend

            return VLLMBackend(
                model=model, base_url=endpoint, api_key=api_key or None
            )
        elif backend_type == "sglang":
            from ollama_classifier.backends import SGLangBackend

            return SGLangBackend(
                model=model, base_url=endpoint, api_key=api_key or None
            )
        elif backend_type == "llamacpp":
            from ollama_classifier.backends import LlamaCppBackend

            return LlamaCppBackend(
                model=model, base_url=endpoint, api_key=api_key or None
            )
        else:
            raise ValueError(f"Unknown backend type: {backend_type}")

    # ==================================================================
    # Settings handlers
    # ==================================================================

    async def _on_backend_type_change(self, e: ft.ControlEvent):
        """Update the default endpoint when the backend type changes."""
        backend = e.control.value
        default_ep = DEFAULT_ENDPOINTS.get(backend, "")
        self.endpoint_field.current.value = default_ep
        self.page.update()

    async def _on_test_connection(self, e: ft.ControlEvent):
        self.test_connection_btn.current.disabled = True
        self.connection_status.current.value = "Testing connection..."
        self.connection_status.current.color = ft.Colors.GREY
        self.page.update()

        backend_type = self.backend_dropdown.current.value
        endpoint = self.endpoint_field.current.value
        api_key = self.api_key_field.current.value or ""

        try:
            if backend_type == "ollama":
                from ollama import AsyncClient

                headers: dict[str, str] = {}
                if api_key:
                    headers["Authorization"] = f"Bearer {api_key}"
                tmp_client = AsyncClient(host=endpoint, headers=headers)
                models_resp = await tmp_client.list()
                model_names = sorted(m.model for m in models_resp.models if m.model)
                if model_names:
                    self.model_field.current.value = model_names[0]
                self.connection_status.current.value = (
                    f"✓ Connected — {len(model_names)} model(s) found."
                )
                self.connection_status.current.color = ft.Colors.GREEN
            else:
                import httpx

                url = f"{endpoint.rstrip('/')}/models"
                resp = httpx.get(
                    url,
                    headers={"Authorization": f"Bearer {api_key}"} if api_key else {},
                    timeout=10.0,
                )
                resp.raise_for_status()
                data = resp.json()
                model_list = data.get("data", [])
                model_names = sorted(m.get("id", m.get("name", "")) for m in model_list)
                if model_names:
                    self.model_field.current.value = model_names[0]
                self.connection_status.current.value = (
                    f"✓ Connected — {len(model_names)} model(s) found."
                )
                self.connection_status.current.color = ft.Colors.GREEN
        except Exception as ex:
            self.connection_status.current.value = f"✗ Error: {ex}"
            self.connection_status.current.color = ft.Colors.RED
        finally:
            self.test_connection_btn.current.disabled = False
            self.page.update()

    async def _on_theme_change(self, e: ft.ControlEvent):
        is_dark = e.control.value
        self.config["theme"] = "dark" if is_dark else "light"
        self.page.theme_mode = ft.ThemeMode.DARK if is_dark else ft.ThemeMode.LIGHT
        save_config(self.config)
        self.page.update()

    async def _on_save_settings(self, e: ft.ControlEvent):
        self.config["backend_type"] = self.backend_dropdown.current.value
        self.config["endpoint"] = self.endpoint_field.current.value
        self.config["model"] = self.model_field.current.value
        self.config["batch_size"] = (
            self.batch_size_field.current.value
            if self.batch_size_field.current
            else None
        ) or "1"
        save_config(self.config)

        api_key = self.api_key_field.current.value
        if api_key:
            await self.secure_storage.set("api_key", api_key)
        else:
            try:
                await self.secure_storage.remove("api_key")
            except Exception:
                pass

        await self._show_dialog(
            "Settings Saved", "Your settings have been saved successfully."
        )

    async def _on_create_shortcut(self, e: ft.ControlEvent):
        if shutil.which("uv") is None:
            await self._show_dialog(
                "uv Not Found",
                "The `uv` package manager is required to create a shortcut.\n\n"
                "Install it from https://docs.astral.sh/uv/getting-started/installation/",
            )
            return
        try:
            from pyshortcuts import make_shortcut

            make_shortcut(
                script="uvx",
                arguments="ollama-classifier-gui",
                name="LLM Classifier GUI",
                terminal=False,
                desktop=True,
                startmenu=True,
            )
            await self._show_dialog(
                "Shortcut Created",
                "Desktop and Start Menu shortcuts have been created.",
            )
        except Exception as ex:
            await self._show_dialog("Error Creating Shortcut", str(ex))

    # ==================================================================
    # Data Input handlers
    # ==================================================================

    async def _on_select_data_file(self, e: ft.ControlEvent):
        import polars as pl

        result = await self._pick_file(
            allowed_extensions=["csv", "xlsx", "xls"],
            dialog_title="Select data file (CSV or Excel)",
        )
        if not result:
            return
        file_path = result
        self.data_file = file_path
        self.data_file_path_text.current.value = file_path
        self.data_sheet_dropdown.current.visible = False
        self.text_column_dropdown.current.visible = False

        try:
            if file_path.endswith((".xlsx", ".xls")):
                sheet_names, self.data_df = self._load_excel_with_sheets(file_path)
                self.data_sheet_dropdown.current.options = [
                    ft.dropdown.Option(n) for n in sheet_names
                ]
                self.data_sheet_dropdown.current.value = sheet_names[0]
                self.data_sheet_dropdown.current.visible = True
            elif file_path.endswith(".csv"):
                self.data_df = pl.read_csv(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_path}")

            if self.data_df is not None and len(self.data_df) > 0:
                cols = self.data_df.columns
                self.text_column_dropdown.current.options = [
                    ft.dropdown.Option(c) for c in cols
                ]
                self.text_column_dropdown.current.value = cols[0]
                self.text_column_dropdown.current.visible = True
                self._refresh_data_preview()
            else:
                raise ValueError("The file contains no data rows.")

        except Exception as ex:
            await self._show_dialog("Error Loading File", str(ex))

        self.page.update()

    async def _on_data_sheet_change(self, e: ft.ControlEvent):
        import polars as pl

        sheet_name = self.data_sheet_dropdown.current.value
        if not sheet_name or not self.data_file:
            return
        try:
            self.data_df = pl.read_excel(
                self.data_file, engine="openpyxl", sheet_name=sheet_name
            )
            cols = self.data_df.columns
            self.text_column_dropdown.current.options = [
                ft.dropdown.Option(c) for c in cols
            ]
            self.text_column_dropdown.current.value = cols[0]
            self._refresh_data_preview()
        except Exception as ex:
            await self._show_dialog("Error Loading Sheet", str(ex))
        self.page.update()

    def _refresh_data_preview(self):
        if self.data_df is None:
            return
        preview = self.data_df.head(5)
        self.data_preview_table.current.columns = [
            ft.DataColumn(ft.Text(c)) for c in preview.columns
        ]
        self.data_preview_table.current.rows = []
        for row in preview.iter_rows(named=True):
            self.data_preview_table.current.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(str(row[c])[:100])) for c in preview.columns
                    ]
                )
            )

    # ==================================================================
    # Schema handlers
    # ==================================================================

    async def _on_add_label(self, e: ft.ControlEvent):
        delete_btn = ft.IconButton(icon=ft.Icons.DELETE)
        row = ft.Row(
            [
                ft.TextField(label="Label", width=200, expand=True),
                ft.TextField(label="Description (optional)", width=300, expand=True),
                delete_btn,
            ],
            spacing=10,
        )
        delete_btn.on_click = lambda ev, r=row: self._on_remove_label(r)
        self.manual_labels_list.current.controls.append(row)
        self.page.update()

    async def _on_remove_label(self, row: ft.Row):
        self.manual_labels_list.current.controls.remove(row)
        self.page.update()

    async def _on_schema_tab_change(self, e: ft.ControlEvent):
        # The Tabs control coordinates TabBar/TabBarView via selected_index;
        # keep it explicit so _collect_labels always reads the current tab.
        self.schema_tabs.current.selected_index = e.control.selected_index
        self.page.update()

    async def _on_select_schema_file(self, e: ft.ControlEvent):
        import polars as pl

        result = await self._pick_file(
            allowed_extensions=["csv", "xlsx", "xls"],
            dialog_title="Select schema / labels file (CSV or Excel)",
        )
        if not result:
            return
        file_path = result
        self.schema_file = file_path
        self.schema_file_path_text.current.value = file_path
        self.schema_sheet_dropdown.current.visible = False

        try:
            if file_path.endswith((".xlsx", ".xls")):
                sheet_names, self.schema_df = self._load_excel_with_sheets(file_path)
                self.schema_sheet_dropdown.current.options = [
                    ft.dropdown.Option(n) for n in sheet_names
                ]
                self.schema_sheet_dropdown.current.value = sheet_names[0]
                self.schema_sheet_dropdown.current.visible = True
            elif file_path.endswith(".csv"):
                self.schema_df = pl.read_csv(file_path)
            else:
                raise ValueError(f"Unsupported file type: {file_path}")

            cols = self.schema_df.columns
            self.label_column_dropdown.current.options = [
                ft.dropdown.Option(c) for c in cols
            ]
            self.description_column_dropdown.current.options = [
                ft.dropdown.Option("-- None --")
            ] + [ft.dropdown.Option(c) for c in cols]
            self.label_column_dropdown.current.value = cols[0]
            self.description_column_dropdown.current.value = "-- None --"
            self.label_column_dropdown.current.visible = True
            self.description_column_dropdown.current.visible = True
            self._refresh_schema_preview()

        except Exception as ex:
            await self._show_dialog("Error Loading Schema File", str(ex))
        self.page.update()

    async def _on_schema_sheet_change(self, e: ft.ControlEvent):
        import polars as pl

        sheet_name = self.schema_sheet_dropdown.current.value
        if not sheet_name or not self.schema_file:
            return
        try:
            self.schema_df = pl.read_excel(
                self.schema_file, engine="openpyxl", sheet_name=sheet_name
            )
            cols = self.schema_df.columns
            self.label_column_dropdown.current.options = [
                ft.dropdown.Option(c) for c in cols
            ]
            self.description_column_dropdown.current.options = [
                ft.dropdown.Option("-- None --")
            ] + [ft.dropdown.Option(c) for c in cols]
            self.label_column_dropdown.current.value = cols[0]
            self.description_column_dropdown.current.value = "-- None --"
            self._refresh_schema_preview()
        except Exception as ex:
            await self._show_dialog("Error Loading Sheet", str(ex))
        self.page.update()

    def _refresh_schema_preview(self):
        if self.schema_df is None:
            return
        label_col = self.label_column_dropdown.current.value
        desc_col = self.description_column_dropdown.current.value
        preview_rows = self.schema_df.head(10)
        lines = []
        for row in preview_rows.iter_rows(named=True):
            label = str(row.get(label_col, ""))
            desc = (
                str(row.get(desc_col, ""))
                if desc_col and desc_col != "-- None --" and desc_col != label_col
                else ""
            )
            lines.append(f"• {label}" + (f" — {desc}" if desc else ""))
        self.schema_preview_text.current.value = (
            "\n".join(lines) if lines else "No labels found."
        )

    def _collect_labels(self) -> dict[str, str]:
        """Collect labels from either manual input or schema file."""
        active_tab = self.schema_tabs.current.selected_index

        if active_tab == 0:
            labels: dict[str, str] = {}
            for row_ctrl in self.manual_labels_list.current.controls:
                fields = [c for c in row_ctrl.controls if isinstance(c, ft.TextField)]
                if len(fields) >= 1 and fields[0].value and fields[0].value.strip():
                    label_name = fields[0].value.strip()
                    desc = (
                        fields[1].value.strip()
                        if len(fields) >= 2 and fields[1].value
                        else ""
                    )
                    labels[label_name] = desc
            return labels
        else:
            if self.schema_df is None:
                return {}
            label_col = self.label_column_dropdown.current.value
            desc_col = self.description_column_dropdown.current.value
            labels = {}
            for row in self.schema_df.iter_rows(named=True):
                name = str(row.get(label_col, "")).strip()
                if not name:
                    continue
                desc = ""
                if desc_col and desc_col != "-- None --" and desc_col != label_col:
                    desc = str(row.get(desc_col, "")).strip()
                labels[name] = desc
            return labels

    # ==================================================================
    # Classification
    # ==================================================================

    async def _on_run_classification(self, e: ft.ControlEvent):
        if self._classifying:
            return

        # --- Validate inputs ---
        if self.data_df is None:
            await self._show_dialog(
                "No Data", "Please load a data file first (Data Input tab)."
            )
            return

        text_col = self.text_column_dropdown.current.value
        if not text_col:
            await self._show_dialog("No Column", "Please select a text column.")
            return

        labels = self._collect_labels()
        if not labels:
            await self._show_dialog(
                "No Labels",
                "Please define at least one classification label (Schema tab).",
            )
            return

        # Build choices
        all_empty_desc = all(d == "" for d in labels.values())
        choices = list(labels.keys()) if all_empty_desc else labels

        method = self.classify_method_radio.current.value
        output_format = self.output_format_radio.current.value
        system_prompt = self.system_prompt_field.current.value or None

        # Parse batch size
        try:
            batch_size = max(1, int(self.batch_size_field.current.value or "1"))
        except ValueError:
            batch_size = 1

        # Initialise classifier
        try:
            classifier = await self._get_classifier()
        except Exception as ex:
            await self._show_dialog(
                "Connection Error", f"Could not initialise classifier:\n{ex}"
            )
            return

        # Extract texts
        texts = self.data_df[text_col].to_list()

        # --- Run ---
        self._classifying = True
        self.run_btn.current.disabled = True
        self.save_btn.current.disabled = True
        self.results_progress.current.visible = True
        self.results_progress.current.value = 0
        self.results_status.current.value = (
            f"Starting classification of {len(texts)} item(s)…"
        )
        self.results = []
        self.results_table.current.rows = []
        self.results_table.current.columns = []
        self.page.update()

        try:
            for i in range(0, len(texts), batch_size):
                batch = texts[i : i + batch_size]
                self.results_status.current.value = (
                    f"Classifying {min(i + batch_size, len(texts))}/{len(texts)}…"
                )
                self.results_progress.current.value = min(
                    i + batch_size, len(texts)
                ) / len(texts)
                self.page.update()

                if batch_size == 1 and method == "generate":
                    # Single-item generation — use per-item async call for progress
                    for text in batch:
                        text_str = str(text)
                        try:
                            if method == "generate":
                                result = await classifier.agenerate(
                                    text=text_str,
                                    choices=choices,
                                    system_prompt=system_prompt,
                                )
                            else:
                                result = await classifier.aclassify(
                                    text=text_str,
                                    choices=choices,
                                    system_prompt=system_prompt,
                                )
                            row_dict = self._result_to_dict(
                                result, text_str, output_format
                            )
                            self.results.append(row_dict)
                            self._add_result_row(
                                row_dict,
                                output_format,
                                i == 0 and len(self.results) == 1,
                            )
                        except Exception as item_ex:
                            row_dict = {
                                "text": text_str,
                                "prediction": f"ERROR: {item_ex}",
                                "confidence": 0.0,
                            }
                            self.results.append(row_dict)
                            self._add_error_row(
                                text_str, i == 0 and len(self.results) == 1
                            )
                else:
                    # Batch classify or batch generate
                    try:
                        if method == "generate":
                            results = await classifier.abatch_generate(
                                texts=[str(t) for t in batch],
                                choices=choices,
                                system_prompt=system_prompt,
                            )
                        else:
                            results = await classifier.abatch_classify(
                                texts=[str(t) for t in batch],
                                choices=choices,
                                system_prompt=system_prompt,
                            )
                        for text_str, result in zip(batch, results):
                            row_dict = self._result_to_dict(
                                result, str(text_str), output_format
                            )
                            self.results.append(row_dict)
                            self._add_result_row(
                                row_dict,
                                output_format,
                                i == 0 and len(self.results) == 1,
                            )
                    except Exception as batch_ex:
                        for text in batch:
                            text_str = str(text)
                            self.results.append(
                                {
                                    "text": text_str,
                                    "prediction": f"ERROR: {batch_ex}",
                                    "confidence": 0.0,
                                }
                            )
                            self._add_error_row(
                                text_str, i == 0 and len(self.results) == 1
                            )

                self.page.update()

            self.results_status.current.value = (
                f"✓ Done — classified {len(texts)} item(s)."
            )
            self.results_status.current.color = ft.Colors.GREEN
            self.save_btn.current.disabled = False

        except Exception as ex:
            self.results_status.current.value = f"✗ Error: {ex}"
            self.results_status.current.color = ft.Colors.RED
        finally:
            self._classifying = False
            self.run_btn.current.disabled = False
            self.results_progress.current.visible = False
            self.page.update()

    def _result_to_dict(
        self, result, text_str: str, output_format: str
    ) -> dict[str, Any]:
        """Convert a ClassificationResult to a dict based on output format."""
        if output_format == "all_labels":
            row_dict: dict[str, Any] = {"text": text_str}
            for label, prob in result.probabilities.items():
                row_dict[label] = prob
            return row_dict
        else:
            return {
                "text": text_str,
                "prediction": result.prediction,
                "confidence": result.confidence,
            }

    def _add_result_row(self, row_dict: dict, output_format: str, is_first: bool):
        """Add a DataRow to the results table."""
        if is_first:
            self.results_table.current.columns = [
                ft.DataColumn(ft.Text(k)) for k in row_dict.keys()
            ]
        cells = []
        for k, v in row_dict.items():
            if isinstance(v, float):
                cells.append(ft.DataCell(ft.Text(f"{v:.2%}")))
            else:
                text = str(v)
                cells.append(
                    ft.DataCell(ft.Text(text[:80] + ("…" if len(text) > 80 else "")))
                )
        self.results_table.current.rows.append(ft.DataRow(cells=cells))

    def _add_error_row(self, text_str: str, is_first: bool):
        """Add an error DataRow."""
        if is_first:
            self.results_table.current.columns = [
                ft.DataColumn(ft.Text("text")),
                ft.DataColumn(ft.Text("prediction")),
                ft.DataColumn(ft.Text("confidence")),
            ]
        self.results_table.current.rows.append(
            ft.DataRow(
                cells=[
                    ft.DataCell(ft.Text(text_str[:80])),
                    ft.DataCell(ft.Text("ERROR", color=ft.Colors.RED)),
                    ft.DataCell(ft.Text("0.00%")),
                ]
            )
        )

    # ==================================================================
    # Save results
    # ==================================================================

    async def _on_save_results(self, e: ft.ControlEvent):
        if not self.results:
            await self._show_dialog(
                "No Results", "Nothing to save — run classification first."
            )
            return

        # Pick save location
        save_path = await self._pick_save_file(
            allowed_extensions=["xlsx"],
            dialog_title="Save results as Excel",
            suggested_name="classification_results.xlsx",
        )
        if not save_path:
            return

        try:
            import polars as pl

            rows = []
            for r in self.results:
                row: dict[str, Any] = {}
                for k, v in r.items():
                    if k == "text":
                        row[k] = v
                    elif k == "prediction":
                        row[k] = v
                    elif k == "confidence":
                        row[k] = v
                    else:
                        # Label probability columns
                        row[k] = v
                rows.append(row)

            results_df = pl.DataFrame(rows)

            # Merge with original data
            text_col = self.text_column_dropdown.current.value
            if text_col and self.data_df is not None:
                # Get original columns and add results columns
                output_format = self.output_format_radio.current.value
                if output_format == "all_labels":
                    extra_cols = [c for c in results_df.columns if c != "text"]
                else:
                    extra_cols = ["prediction", "confidence"]

                # Build output: original data + results columns aligned by row index
                original_cols = self.data_df.columns
                output_data: dict[str, list] = {}
                for col in original_cols:
                    output_data[col] = self.data_df[col].to_list()

                n_texts = len(self.data_df)
                n_results = len(self.results)
                for col in extra_cols:
                    col_values = []
                    for idx in range(n_texts):
                        if idx < n_results and col in self.results[idx]:
                            col_values.append(self.results[idx][col])
                        else:
                            col_values.append(None)
                    output_data[col] = col_values

                output_df = pl.DataFrame(output_data)
            else:
                output_df = results_df

            output_df.write_excel(save_path)
            await self._show_dialog("Results Saved", f"Results saved to:\n{save_path}")
        except Exception as ex:
            await self._show_dialog("Error Saving Results", str(ex))

    # ==================================================================
    # Helpers
    # ==================================================================

    @staticmethod
    def _load_excel_with_sheets(path: str) -> tuple[list[str], Any]:
        """Discover an Excel workbook's sheet names and read its first sheet.

        Guards against workbooks with no sheets (raises a clear ``ValueError``)
        so both the data and schema loaders share identical edge-case handling.
        Returns ``(sheet_names, first_sheet_df)``.
        """
        import openpyxl
        import polars as pl

        wb = openpyxl.load_workbook(path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        if not sheet_names:
            raise ValueError("Excel file has no sheets.")
        df = pl.read_excel(path, engine="openpyxl", sheet_name=sheet_names[0])
        return sheet_names, df

    async def _pick_file(
        self,
        allowed_extensions: list[str] | None = None,
        dialog_title: str = "Select file",
    ) -> str | None:
        files = await ft.FilePicker().pick_files(
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=allowed_extensions or [],
            dialog_title=dialog_title,
        )
        if files:
            return files[0].path
        return None

    async def _pick_save_file(
        self,
        allowed_extensions: list[str] | None = None,
        dialog_title: str = "Save file",
        suggested_name: str = "results.xlsx",
    ) -> str | None:
        return await ft.FilePicker().save_file(
            file_type=ft.FilePickerFileType.CUSTOM,
            allowed_extensions=allowed_extensions or [],
            dialog_title=dialog_title,
            file_name=suggested_name,
        )

    async def _show_dialog(self, title: str, message: str):
        dismissed = asyncio.Event()

        def _on_dismiss(e):
            dismissed.set()

        dlg = ft.AlertDialog(
            title=ft.Text(title),
            content=ft.Text(message, selectable=True),
            actions=[ft.TextButton("OK", on_click=lambda e: self.page.pop_dialog())],
            on_dismiss=_on_dismiss,
        )
        self.page.show_dialog(dlg)
        await dismissed.wait()


# ======================================================================
# Entry point
# ======================================================================


async def _flet_main(page: ft.Page):
    """Flet entry point — creates the app and runs its main loop."""
    app = OllamaClassifierApp(page)
    await app.main()


def main():
    """Entry point for the Flet app (desktop mode)."""
    ft.run(main=_flet_main)


if __name__ == "__main__":
    main()
